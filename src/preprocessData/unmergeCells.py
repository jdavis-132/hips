import openpyxl
import csv
from pathlib import Path

# Load the workbook
file_path = 'rawData/2023/inbreds/2023_Inbred_Hips_Ear_Phenotyping_MV_LNK_Final_KL_Curated.xlsx'
wb = openpyxl.load_workbook(file_path)

# Get the 'Data Entry Sheet'
sheet = wb['Data Entry Sheet']

# Prepare data structure to store unmerged data
data = []
for row in sheet.iter_rows():
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    data.append(row_data)

# Get merged cell ranges
merged_ranges = list(sheet.merged_cells.ranges)

# For each merged cell range in columns A-J (0-9)
for merged_range in merged_ranges:
    # Get the top-left cell value
    min_col = merged_range.min_col
    min_row = merged_range.min_row
    max_col = merged_range.max_col
    max_row = merged_range.max_row

    # Only process if any part of the merge is in columns A-J (1-10)
    if min_col <= 10:
        # Get the value from the top-left cell
        value = data[min_row - 1][min_col - 1]

        # Fill all cells in the merged range (but only for columns A-J)
        for row_idx in range(min_row - 1, max_row):
            for col_idx in range(min_col - 1, min(max_col, 10)):
                data[row_idx][col_idx] = value

# Write to CSV
output_path = 'rawData/2023_Inbred_Hips_Ear_Phenotyping_MV_LNK_Final_KL_Curated_DataEntrySheet.csv'
with open(output_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerows(data)

print(f"Successfully created: {output_path}")
print(f"Total rows: {len(data)}")
print(f"Total columns: {len(data[0]) if data else 0}")
print(f"Processed {len(merged_ranges)} merged cell ranges")
